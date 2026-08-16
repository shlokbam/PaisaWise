from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, desc
from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.user import User
from app.models.transaction import Transaction
from app.models.feedback import UserFeedback
from app.models.rule import Rule
from app.models.merchant import Merchant
from app.models.notification import Notification
from app.schemas.transaction import TransactionOut, TransactionCreate, TransactionUpdate, TransactionFeedback
from app.ai.classifier import classify_transaction_with_ai
from typing import List, Optional
from decimal import Decimal
from datetime import date, timedelta, datetime
import csv
import io
from fpdf import FPDF

router = APIRouter(prefix="/transactions", tags=["transactions"])

@router.get("", response_model=List[TransactionOut])
def get_transactions(
    ownership: Optional[str] = None,
    type: Optional[str] = None,
    include: Optional[bool] = None,
    needs_review: Optional[bool] = None,
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    List and filter user transactions.
    Supports search (merchant name, upi id, description) and preset filters.
    """
    query = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    
    if category_id:
        query = query.filter(Transaction.category_id == category_id)
    if ownership:
        query = query.filter(Transaction.ownership == ownership.upper())
    if type:
        query = query.filter(Transaction.transaction_type == type.upper())
    if include is not None:
        query = query.filter(Transaction.include_in_personal_expenses == include)
    if needs_review is not None and needs_review:
        query = query.filter(Transaction.confidence < Decimal("0.90"))
        
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                Transaction.merchant_name.ilike(search_filter),
                Transaction.upi_id.ilike(search_filter),
                Transaction.description.ilike(search_filter)
            )
        )
        
    # Sort by date desc, time desc
    query = query.order_by(desc(Transaction.transaction_date), desc(Transaction.transaction_time))
    
    # Paginate
    offset = (page - 1) * limit
    return query.offset(offset).limit(limit).all()

@router.get("/export")
def export_transactions(
    format: str = "csv",
    range_type: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export transaction list to CSV or PDF based on date range criteria.
    Format options: 'csv', 'pdf'
    Range type options: 'week', 'month', 'year', 'custom'
    """
    today = date.today()
    query_start = None
    query_end = today + timedelta(days=1) # include today
    
    if range_type == "week":
        query_start = today - timedelta(days=7)
    elif range_type == "month":
        query_start = today - timedelta(days=30)
    elif range_type == "year":
        query_start = today - timedelta(days=365)
    elif range_type == "custom":
        if start_date:
            query_start = datetime.fromisoformat(start_date).date()
        if end_date:
            query_end = datetime.fromisoformat(end_date).date() + timedelta(days=1)
            
    query = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    if query_start:
        query = query.filter(Transaction.transaction_date >= query_start)
    if query_end:
        query = query.filter(Transaction.transaction_date < query_end)
        
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    if format == "csv":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ledger Report"
        
        # Grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Colors
        PURPLE_HEADER = "4F46E5"
        WHITE = "FFFFFF"
        LIGHT_SLATE = "F1F5F9"
        GREEN_TEXT = "10B981"
        RED_TEXT = "F43F5E"
        DARK_TEXT = "0F172A"
        MUTED_TEXT = "475569"
        
        # Styles
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
        font_regular = Font(name="Calibri", size=11, color=DARK_TEXT)
        font_muted = Font(name="Calibri", size=10, italic=True, color=MUTED_TEXT)
        
        fill_title = PatternFill(start_color="0B0F19", end_color="0B0F19", fill_type="solid")
        fill_header = PatternFill(start_color=PURPLE_HEADER, end_color=PURPLE_HEADER, fill_type="solid")
        fill_alt = PatternFill(start_color=LIGHT_SLATE, end_color=LIGHT_SLATE, fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color="CBD5E1"),
            right=Side(style='thin', color="CBD5E1"),
            top=Side(style='thin', color="CBD5E1"),
            bottom=Side(style='thin', color="CBD5E1")
        )
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # Header Banner
        ws.merge_cells("A1:I2")
        title_cell = ws["A1"]
        title_cell.value = "PaisaWise Financial Statement"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Statement Metadata (Row 4-6)
        ws["A4"] = "Generated on:"
        ws["A4"].font = font_bold
        ws["B4"] = today.isoformat()
        ws["B4"].font = font_regular
        
        ws["A5"] = "Scope:"
        ws["A5"].font = font_bold
        ws["B5"] = f"{range_type.upper()} ({query_start or 'All'} to {query_end or 'Today'})"
        ws["B5"].font = font_regular
        
        ws["A6"] = "User Account:"
        ws["A6"].font = font_bold
        ws["B6"] = current_user.email
        ws["B6"].font = font_regular
        
        # Summary Box Cards (D4:G6)
        total_inflow = sum(tx.amount for tx in transactions if tx.transaction_type == "INCOME")
        total_outflow = sum(tx.amount for tx in transactions if tx.transaction_type == "EXPENSE")
        personal_spending = sum(tx.amount for tx in transactions if tx.include_in_personal_expenses)
        investments = sum(tx.amount for tx in transactions if tx.transaction_type == "INVESTMENT")
        
        ws["D4"] = "Total Inflow"
        ws["D4"].font = font_muted
        ws["D5"] = total_inflow
        ws["D5"].font = Font(name="Calibri", size=13, bold=True, color=GREEN_TEXT)
        ws["D5"].number_format = '"INR "#,##0.00'
        
        ws["E4"] = "Total Outflow"
        ws["E4"].font = font_muted
        ws["E5"] = total_outflow
        ws["E5"].font = Font(name="Calibri", size=13, bold=True, color=RED_TEXT)
        ws["E5"].number_format = '"INR "#,##0.00'
        
        ws["F4"] = "Personal Spending"
        ws["F4"].font = font_muted
        ws["F5"] = personal_spending
        ws["F5"].font = Font(name="Calibri", size=13, bold=True, color=PURPLE_HEADER)
        ws["F5"].number_format = '"INR "#,##0.00'

        ws["G4"] = "Investments"
        ws["G4"].font = font_muted
        ws["G5"] = investments
        ws["G5"].font = Font(name="Calibri", size=13, bold=True, color="2563EB")
        ws["G5"].number_format = '"INR "#,##0.00'
        
        # Header Row (Row 8)
        headers = [
            "Date", "Merchant/Sender", "Amount (INR)", "Type", "Ownership", 
            "Category", "Payment Method", "Status", "Confidence"
        ]
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header_title
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            
        # Data Rows (Row 9+)
        row_idx = 9
        for tx in transactions:
            category_name = tx.category.name if tx.category else "Uncategorized"
            status = "Personal Spending" if tx.include_in_personal_expenses else "Excluded"
            confidence_val = (tx.confidence if tx.confidence is not None else Decimal("1.00"))
            
            row_data = [
                tx.transaction_date,
                tx.merchant_name or tx.sender or "Unknown",
                tx.amount,
                tx.transaction_type,
                tx.ownership,
                category_name,
                tx.payment_method or "UPI",
                status,
                confidence_val
            ]
            
            fill_row = fill_alt if row_idx % 2 == 0 else None
            
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.value = val
                cell.font = font_regular
                cell.border = border_thin
                if fill_row:
                    cell.fill = fill_row
                    
                # Format cell data
                if col_num == 1: # Date
                    cell.number_format = "YYYY-MM-DD"
                    cell.alignment = align_center
                elif col_num == 2 or col_num == 6: # Merchant / Category
                    cell.alignment = align_left
                elif col_num == 3: # Amount
                    cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif col_num == 4: # Type (Color dynamically!)
                    cell.alignment = align_center
                    if val == "INCOME":
                        cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN_TEXT)
                    elif val == "EXPENSE":
                        cell.font = Font(name="Calibri", size=11, bold=True, color=RED_TEXT)
                elif col_num == 9: # Confidence
                    cell.number_format = "0%"
                    cell.alignment = align_center
                else:
                    cell.alignment = align_center
                    
            row_idx += 1
            
        # Auto-fit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                if cell.value:
                    if isinstance(cell.value, (int, float, Decimal)):
                        max_len = max(max_len, 12)
                    else:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"paisawise_report_{range_type}_{today.isoformat()}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    elif format == "pdf":
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_margins(10, 10, 10)
        
        # Draw Header Banner (Deep Blue background)
        pdf.set_fill_color(11, 15, 25) # dark.bg
        pdf.rect(0, 0, 210, 40, "F")
        
        # PaisaWise Logo / Title
        pdf.set_font("helvetica", "B", 18)
        pdf.set_text_color(248, 250, 252) # dark.text
        pdf.set_xy(10, 12)
        pdf.cell(0, 8, "PaisaWise Financial Statement", align="L")
        
        # Sub-header Info (Right-aligned in Header Banner)
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(148, 163, 184) # dark.muted
        pdf.set_xy(120, 10)
        pdf.cell(80, 5, f"Generated: {today.isoformat()}", ln=True, align="R")
        pdf.set_xy(120, 15)
        pdf.cell(80, 5, f"Account: {current_user.email}", ln=True, align="R")
        pdf.set_xy(120, 20)
        pdf.cell(80, 5, f"Scope: {range_type.upper()}", ln=True, align="R")
        
        # Add a colored divider bar below header banner
        pdf.set_fill_color(79, 70, 229) # dark.accent (purple)
        pdf.rect(0, 40, 210, 2, "F")
        
        pdf.set_xy(10, 48)
        
        # Summary Box Cards (Styled grid)
        total_inflow = sum(tx.amount for tx in transactions if tx.transaction_type == "INCOME")
        total_outflow = sum(tx.amount for tx in transactions if tx.transaction_type == "EXPENSE")
        personal_spending = sum(tx.amount for tx in transactions if tx.include_in_personal_expenses)
        investments = sum(tx.amount for tx in transactions if tx.transaction_type == "INVESTMENT")
        
        pdf.set_font("helvetica", "B", 12)
        pdf.set_text_color(11, 15, 25)
        pdf.cell(0, 8, "Financial Summary Metrics", ln=True)
        pdf.ln(2)
        
        # Render 4 Cards
        # Total inflow Card (Green)
        pdf.set_fill_color(240, 253, 250) # Light teal bg
        pdf.set_draw_color(16, 185, 129) # Teal border
        pdf.rect(10, 58, 44, 18, "DF")
        pdf.set_xy(12, 60)
        pdf.set_font("helvetica", "", 8)
        pdf.set_text_color(71, 85, 105)
        pdf.cell(40, 4, "Total Inflow", ln=True)
        pdf.set_xy(12, 65)
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(16, 185, 129)
        pdf.cell(40, 5, f"INR {total_inflow:,.2f}")
        
        # Total Outflow Card (Red/Rose)
        pdf.set_fill_color(255, 241, 242) # Light rose bg
        pdf.set_draw_color(244, 63, 94) # Rose border
        pdf.rect(57, 58, 44, 18, "DF")
        pdf.set_xy(59, 60)
        pdf.set_font("helvetica", "", 8)
        pdf.set_text_color(71, 85, 105)
        pdf.cell(40, 4, "Total Outflow", ln=True)
        pdf.set_xy(59, 65)
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(244, 63, 94)
        pdf.cell(40, 5, f"INR {total_outflow:,.2f}")
        
        # Personal Spending Card (Purple)
        pdf.set_fill_color(245, 243, 255) # Light purple bg
        pdf.set_draw_color(79, 70, 229) # Purple border
        pdf.rect(104, 58, 44, 18, "DF")
        pdf.set_xy(106, 60)
        pdf.set_font("helvetica", "", 8)
        pdf.set_text_color(71, 85, 105)
        pdf.cell(40, 4, "Personal Spending", ln=True)
        pdf.set_xy(106, 65)
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(79, 70, 229)
        pdf.cell(40, 5, f"INR {personal_spending:,.2f}")
        
        # Investments Card (Blue)
        pdf.set_fill_color(239, 246, 255) # Light blue bg
        pdf.set_draw_color(37, 99, 235) # Blue border
        pdf.rect(151, 58, 49, 18, "DF")
        pdf.set_xy(153, 60)
        pdf.set_font("helvetica", "", 8)
        pdf.set_text_color(71, 85, 105)
        pdf.cell(45, 4, "Investments", ln=True)
        pdf.set_xy(153, 65)
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(37, 99, 235)
        pdf.cell(45, 5, f"INR {investments:,.2f}")
        
        pdf.set_xy(10, 82)
        pdf.set_font("helvetica", "B", 12)
        pdf.set_text_color(11, 15, 25)
        pdf.cell(0, 8, "Transaction Ledger", ln=True)
        pdf.ln(2)
        
        # Table Header
        widths = [20, 45, 28, 18, 35, 18, 16, 10]
        headers = ["Date", "Merchant/Sender", "Amount (INR)", "Type", "Category", "Payment", "Pers.?", "Conf"]
        
        pdf.set_fill_color(79, 70, 229) # Purple header bg
        pdf.set_draw_color(203, 213, 225)  # Slate border
        pdf.set_font("helvetica", "B", 8)
        pdf.set_text_color(255, 255, 255)
        
        for w, h_title in zip(widths, headers):
            pdf.cell(w, 8, h_title, border=1, align="C", fill=True)
        pdf.ln()
        
        # Table Data Rows
        pdf.set_font("helvetica", "", 8)
        pdf.set_text_color(15, 23, 42)
        
        row_idx = 0
        for tx in transactions:
            category_name = tx.category.name if tx.category else "Uncategorized"
            is_personal = "Yes" if tx.include_in_personal_expenses else "No"
            merchant = tx.merchant_name or tx.sender or "Unknown"
            confidence_pct = f"{int(tx.confidence * 100)}%" if tx.confidence is not None else "100%"
            
            if len(merchant) > 24:
                merchant = merchant[:21] + "..."
                
            if row_idx % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(241, 245, 249) # Light slate grey
                
            pdf.cell(widths[0], 8, tx.transaction_date.isoformat(), border=1, align="C", fill=True)
            pdf.cell(widths[1], 8, merchant, border=1, fill=True)
            
            # Amount
            pdf.set_font("helvetica", "B", 8)
            if tx.transaction_type == "EXPENSE":
                pdf.set_text_color(244, 63, 94) # Rose
            elif tx.transaction_type == "INCOME":
                pdf.set_text_color(16, 185, 129) # Green
            else:
                pdf.set_text_color(15, 23, 42)
            pdf.cell(widths[2], 8, f"{tx.amount:,.2f}", border=1, align="R", fill=True)
            
            pdf.set_font("helvetica", "", 8)
            pdf.set_text_color(15, 23, 42)
            
            pdf.cell(widths[3], 8, tx.transaction_type, border=1, align="C", fill=True)
            pdf.cell(widths[4], 8, category_name, border=1, fill=True)
            pdf.cell(widths[5], 8, tx.payment_method or "UPI", border=1, align="C", fill=True)
            pdf.cell(widths[6], 8, is_personal, border=1, align="C", fill=True)
            pdf.cell(widths[7], 8, confidence_pct, border=1, align="C", fill=True)
            pdf.ln()
            row_idx += 1
            
        filename = f"paisawise_statement_{range_type}_{today.isoformat()}.pdf"
        pdf_bytes = pdf.output()
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    raise HTTPException(status_code=400, detail="Invalid export format. Must be 'csv' or 'pdf'.")

@router.get("/{id}", response_model=TransactionOut)
def get_transaction(
    id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Fetch a single transaction detail."""
    tx = db.query(Transaction).filter(Transaction.id == id, Transaction.user_id == current_user.id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return tx

@router.post("", response_model=TransactionOut)
def create_transaction(
    tx_in: TransactionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Manually record a transaction."""
    tx = Transaction(
        user_id=current_user.id,
        account_id=tx_in.account_id,
        amount=tx_in.amount,
        currency=tx_in.currency,
        direction=tx_in.direction.upper(),
        transaction_date=tx_in.transaction_date,
        transaction_time=tx_in.transaction_time,
        merchant_name=tx_in.merchant_name,
        upi_id=tx_in.upi_id,
        sender=tx_in.sender,
        receiver=tx_in.receiver,
        payment_method=tx_in.payment_method.upper(),
        description=tx_in.description,
        ownership=tx_in.ownership.upper(),
        transaction_type=tx_in.transaction_type.upper(),
        category_id=tx_in.category_id,
        subcategory_id=tx_in.subcategory_id,
        confidence=Decimal("1.00"), # Manually entered = 100% confidence
        include_in_personal_expenses=tx_in.include_in_personal_expenses,
        source="MANUAL",
        is_duplicate=False
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)
    return tx

@router.patch("/{id}", response_model=TransactionOut)
def update_transaction(
    id: str,
    tx_in: TransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Manually update/edit a transaction. Sets confidence to 1.0."""
    tx = db.query(Transaction).filter(Transaction.id == id, Transaction.user_id == current_user.id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    update_data = tx_in.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(tx, field, value)
        
    tx.confidence = Decimal("1.00") # Manually adjusted
    
    # Calculate include_in_personal_expenses automatically unless explicitly overridden
    if "include_in_personal_expenses" not in update_data:
        if tx.ownership == "PERSONAL" and tx.transaction_type == "EXPENSE":
            tx.include_in_personal_expenses = True
        else:
            tx.include_in_personal_expenses = False
            
    db.commit()
    db.refresh(tx)
    return tx

@router.post("/{id}/feedback")
def submit_feedback(
    id: str,
    feedback_in: TransactionFeedback,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Submits user correction/confirmation for AI classification.
    Learns from feedback and suggests rule generation if patterns recur.
    """
    tx = db.query(Transaction).filter(Transaction.id == id, Transaction.user_id == current_user.id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    # Save correction logs
    feedback = db.query(UserFeedback).filter_by(transaction_id=tx.id).first()
    if not feedback:
        feedback = UserFeedback(user_id=current_user.id, transaction_id=tx.id)
        db.add(feedback)
        
    # Record corrections
    if feedback_in.ownership:
        feedback.corrected_ownership = feedback_in.ownership.upper()
        tx.ownership = feedback_in.ownership.upper()
    if feedback_in.transaction_type:
        feedback.corrected_type = feedback_in.transaction_type.upper()
        tx.transaction_type = feedback_in.transaction_type.upper()
    if feedback_in.category_id:
        feedback.corrected_category_id = feedback_in.category_id
        tx.category_id = feedback_in.category_id
    if feedback_in.subcategory_id:
        tx.subcategory_id = feedback_in.subcategory_id
    if feedback_in.include_in_personal_expenses is not None:
        feedback.corrected_include = feedback_in.include_in_personal_expenses
        tx.include_in_personal_expenses = feedback_in.include_in_personal_expenses
    else:
        # Determine inclusion rule
        if tx.ownership == "PERSONAL" and tx.transaction_type == "EXPENSE":
            tx.include_in_personal_expenses = True
        else:
            tx.include_in_personal_expenses = False
            
    tx.confidence = Decimal("1.00") # Confirmed by human-in-the-loop
    db.commit()
    db.refresh(tx)
    
    # Clean up corresponding notification
    notif = db.query(Notification).filter_by(transaction_id=tx.id, status="PENDING").first()
    if notif:
        notif.status = "READ"
        db.commit()
        
    # check for Rule Suggestion
    # Let's count how many similar corrections have been made for this merchant/UPI pattern
    suggest_rule = False
    rule_suggestion_payload = {}
    
    if tx.merchant_name or tx.upi_id:
        filter_clause = []
        if tx.merchant_name:
            filter_clause.append(Transaction.merchant_name == tx.merchant_name)
        if tx.upi_id:
            filter_clause.append(Transaction.upi_id == tx.upi_id)
            
        similar_corrections_count = (
            db.query(UserFeedback)
            .join(Transaction, Transaction.id == UserFeedback.transaction_id)
            .filter(
                UserFeedback.user_id == current_user.id,
                or_(*filter_clause)
            )
            .count()
        )
        
        # Suggest a rule if we have corrected this merchant/UPI at least 2 times
        if similar_corrections_count >= 2:
            # Check if a rule already exists for this pattern
            existing_rule = db.query(Rule).filter(
                Rule.user_id == current_user.id,
                or_(
                    and_(Rule.merchant_pattern == tx.merchant_name, Rule.merchant_pattern != None),
                    and_(Rule.upi_pattern == tx.upi_id, Rule.upi_pattern != None)
                )
            ).first()
            
            if not existing_rule:
                suggest_rule = True
                rule_suggestion_payload = {
                    "name": f"Rule for {tx.merchant_name or tx.upi_id}",
                    "merchant_pattern": tx.merchant_name,
                    "upi_pattern": tx.upi_id,
                    "set_ownership": tx.ownership,
                    "set_transaction_type": tx.transaction_type,
                    "set_category_id": tx.category_id,
                    "set_subcategory_id": tx.subcategory_id,
                    "set_include_in_personal_expenses": tx.include_in_personal_expenses
                }
                
    return {
        "status": "success",
        "transaction": tx,
        "suggest_rule": suggest_rule,
        "rule_suggestion": rule_suggestion_payload
    }

@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_transaction(
    id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Deletes a transaction permanently."""
    tx = db.query(Transaction).filter(Transaction.id == id, Transaction.user_id == current_user.id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    db.query(Notification).filter_by(transaction_id=tx.id).delete()
    db.query(UserFeedback).filter_by(transaction_id=tx.id).delete()
    db.delete(tx)
    db.commit()
    return None
